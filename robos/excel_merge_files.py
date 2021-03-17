from openpyxl import load_workbook, Workbook

# lista de arquivos
lista_arquivos = ['CustosAutom', 'PopulacaoPOA', 'SuperMercado']

# Novo arquivo

wb = Workbook()
nome_arquivo_final = "resultado.xlsx"

for nome in lista_arquivos:
    arquivo = load_workbook(filename="%s.xlsx" % nome)
    sheet = arquivo[nome]